#coding=utf-8
#!/usr/bin/python
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import json
import requests


# Read JSON data
json_data = 'null'
#计数
flag = 0

# 读取Excel表格数据
workbook = openpyxl.load_workbook(r'./comp.xlsx')
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    cell_value = row[0]

    # 发送GET请求
    url = 'https://open.api.tianyancha.com/services/open/ic/baseinfoV2/2.0'
    params = {'keyword': cell_value}

    ###↓↓↓↓↓
    headers = {'Authorization': '此处填写你的token'}
    ###↑↑↑↑↑

    response = requests.get(url, params=params, headers=headers)

    # 处理响应
    if response.status_code == 200:
        # 在这里处理响应的数据
        json_data=response.json()
        flag = flag +1
        print("当前次数：" + str(flag) + "  " +cell_value+" 请求成功\n")
    else:
        print('请求失败')


    #失败监测
    if json_data["error_code"] !=0:
        # Open or create the Excel file
        try:
            workbook_err = openpyxl.load_workbook(r"./error.xlsx")
        except FileNotFoundError:
            workbook_err = openpyxl.Workbook()

        # Select the active sheet
        sheet_err = workbook_err.active

        # Determine the next available row in the worksheet
        next_row_err = sheet_err.max_row + 1
        sheet_err.cell(row=next_row_err, column=1).value = cell_value
        # Save the Excel file
        workbook_err.save(r"./error.xlsx")
        print(cell_value+" 获取失败----------------\n")
        continue


    # 写入Excel文件
    # 读 需要加上header=None
    df = pd.read_excel('./example.xlsx',header=None)


    # 写 打开现有的Excel文件
    workbook_out = openpyxl.load_workbook('./output.xlsx')
    # 选择要操作的工作表
    worksheet_out = workbook_out.active
    # 定义要追加的信息列表
    output_list = []


    ss = df.head()
    # 将dataframe类型转化为list类型
    info_list = ss.values.tolist()


    #json处理
    #data=json.loads(json_data)
    data=json_data

    # 打印第1行的内容,注意python中list从0开始计数
    for value in info_list[0]:
        try:
            value_str = str(data['result'][value])
            #print(value + "  " + str(data['result'][value]))
        except KeyError:
            value_str = "  ---"
            #print(value + "  ---")

        output_list.append(value_str)


    # 追加信息到Excel文件的下一行
    worksheet_out.append(output_list)
    # 保存修改后的Excel文件
    workbook_out.save('./output.xlsx')
